import argparse
import colorsys
import math
import os
from pathlib import Path

os.environ["MPLCONFIGDIR"] = "/tmp/matplotlib"

import matplotlib
matplotlib.use("Agg")

import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
from openpyxl import load_workbook


DEFAULT_INPUT = Path(
    r"C:\Users\megamind\Desktop\contribution-matrices.xlsx"
)
DEFAULT_SHEET = "contribution_matrix"
DEFAULT_OUTPUT = Path(
    r"C:\Users\megamind\Desktop\contribution-matrix-visualization.png"
)
DEFAULT_EDGE_PERCENTILE = 88.0
DEFAULT_TOP_EDGES_PER_NODE = 4
DEFAULT_LAYOUT_SEED = 42
DEFAULT_LAYOUT_K = 1.05
DEFAULT_LAYOUT_SCALE = 0.72
DEFAULT_FONT_SIZE = 14
DEFAULT_PERIPHERAL_PULL = 0.68


def read_matrix(file_path: Path, sheet_name: str) -> tuple[list[str], np.ndarray]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))

    labels = [str(cell).strip() for cell in rows[0][1:] if cell is not None]
    size = len(labels)
    matrix = np.zeros((size, size), dtype=float)

    for row_idx, row in enumerate(rows[1:]):
        values = row[1 : size + 1]
        for col_idx, value in enumerate(values):
            matrix[row_idx, col_idx] = float(value or 0)

    return labels, matrix


def build_graph(
    labels: list[str],
    matrix: np.ndarray,
    edge_percentile: float,
    top_edges_per_node: int,
) -> nx.Graph:
    graph = nx.Graph()
    diagonal = np.diag(matrix)

    for label, node_weight in zip(labels, diagonal):
        graph.add_node(label, weight=float(node_weight))

    edge_weights = []
    edge_candidates = []
    for i, source in enumerate(labels):
        for j in range(i + 1, len(labels)):
            target = labels[j]
            weight = float(matrix[i, j])
            if weight > 0:
                edge_weights.append(weight)
                edge_candidates.append((source, target, weight))

    if not edge_candidates:
        raise ValueError("矩阵里没有可视化的非零边。")

    threshold = np.percentile(edge_weights, edge_percentile)
    selected_edges = set()

    for index, label in enumerate(labels):
        neighbors = []
        for other_index, other_label in enumerate(labels):
            if index == other_index:
                continue
            weight = float(matrix[index, other_index])
            if weight > 0:
                neighbors.append((weight, tuple(sorted((label, other_label)))))

        neighbors.sort(reverse=True)
        for _, edge_key in neighbors[:top_edges_per_node]:
            selected_edges.add(edge_key)

    for source, target, weight in edge_candidates:
        edge_key = tuple(sorted((source, target)))
        if weight >= threshold and edge_key in selected_edges:
            graph.add_edge(source, target, weight=weight)

    # 保证每个节点至少保留一条最强边，避免孤点影响布局。
    for label in labels:
        if graph.degree(label) > 0:
            continue

        idx = labels.index(label)
        best_weight = 0.0
        best_neighbor = None
        for j, other in enumerate(labels):
            if idx == j:
                continue
            weight = float(matrix[idx, j])
            if weight > best_weight:
                best_weight = weight
                best_neighbor = other

        if best_neighbor is not None and best_weight > 0:
            graph.add_edge(label, best_neighbor, weight=best_weight)

    return graph


def assign_unique_colors(labels: list[str]) -> dict[str, str]:
    color_map = {}
    total = max(len(labels), 1)
    for index, label in enumerate(labels):
        hue = index / total
        saturation = 0.68
        value = 0.95
        red, green, blue = colorsys.hsv_to_rgb(hue, saturation, value)
        color_map[label] = "#{:02x}{:02x}{:02x}".format(
            int(red * 255),
            int(green * 255),
            int(blue * 255),
        )
    return color_map


def scale_values(values: list[float], low: float, high: float) -> list[float]:
    if not values:
        return []
    minimum = min(values)
    maximum = max(values)
    if math.isclose(minimum, maximum):
        return [(low + high) / 2 for _ in values]
    return [
        low + (value - minimum) * (high - low) / (maximum - minimum)
        for value in values
    ]


def pull_peripheral_nodes_inward(graph: nx.Graph, positions: dict[str, np.ndarray], factor: float) -> dict[str, np.ndarray]:
    center = np.mean(np.array(list(positions.values())), axis=0)
    adjusted = {}
    for node, pos in positions.items():
        offset = np.array(pos) - center
        if graph.degree(node) <= 1:
            adjusted[node] = center + offset * factor
        else:
            adjusted[node] = np.array(pos)
    return adjusted


def draw_graph(
    graph: nx.Graph,
    output_path: Path,
    title: str | None,
    seed: int,
    layout_k: float,
    layout_scale: float,
    font_size: int,
    peripheral_pull: float,
) -> None:
    ordered_nodes = list(graph.nodes)
    color_map = assign_unique_colors(ordered_nodes)
    positions = nx.spring_layout(
        graph,
        weight="weight",
        k=layout_k,
        scale=layout_scale,
        iterations=500,
        seed=seed,
    )
    positions = pull_peripheral_nodes_inward(graph, positions, peripheral_pull)

    node_weights = [graph.nodes[node]["weight"] for node in ordered_nodes]
    node_sizes = scale_values(node_weights, 140, 1300)
    node_colors = [color_map[node] for node in ordered_nodes]

    edges = list(graph.edges(data=True))
    edge_weights = [edge_data["weight"] for _, _, edge_data in edges]
    edge_widths = scale_values(edge_weights, 1.0, 8.5)
    edge_alphas = scale_values(edge_weights, 0.18, 0.78)

    fig, ax = plt.subplots(figsize=(16, 12), facecolor="white")
    ax.set_facecolor("white")

    for (source, target, _), width, alpha in zip(edges, edge_widths, edge_alphas):
        nx.draw_networkx_edges(
            graph,
            positions,
            edgelist=[(source, target)],
            ax=ax,
            width=width,
            edge_color="#6f8fb0",
            alpha=alpha,
        )
    nx.draw_networkx_nodes(
        graph,
        positions,
        ax=ax,
        node_size=node_sizes,
        node_color=node_colors,
        linewidths=0.0,
    )

    label_positions = {}
    center = np.mean(np.array(list(positions.values())), axis=0)
    for node, (x, y) in positions.items():
        direction = np.array([x - center[0], y - center[1]])
        norm = np.linalg.norm(direction)
        if norm == 0:
            direction = np.array([0.012, 0.012])
        else:
            direction = direction / norm * 0.02
        label_positions[node] = (x + direction[0], y + direction[1])

    labels = nx.draw_networkx_labels(
        graph,
        label_positions,
        ax=ax,
        font_size=font_size,
        font_color="#222222",
        font_family="DejaVu Sans",
        font_weight="bold",
    )
    for text in labels.values():
        text.set_bbox(
            dict(
                facecolor="white",
                edgecolor="#bdbdbd",
                boxstyle="square,pad=0.12",
                linewidth=0.8,
            )
        )

    if title:
        ax.set_title(title, fontsize=18, color="#222222", pad=16)

    ax.set_axis_off()
    fig.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=220, facecolor=fig.get_facecolor(), bbox_inches="tight")
    plt.close(fig)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="把 contribution-matrices.xlsx 可视化为关系网络图")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="输入矩阵 Excel 路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="使用的工作表名称")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出图片路径")
    parser.add_argument(
        "--edge-percentile",
        type=float,
        default=DEFAULT_EDGE_PERCENTILE,
        help="仅保留高于该分位数的边，范围 0-100，默认 88",
    )
    parser.add_argument(
        "--top-edges-per-node",
        type=int,
        default=DEFAULT_TOP_EDGES_PER_NODE,
        help="每个词最多保留的核心连边数，默认 4",
    )
    parser.add_argument("--layout-k", type=float, default=DEFAULT_LAYOUT_K, help="布局紧凑度，越小越紧")
    parser.add_argument(
        "--layout-scale",
        type=float,
        default=DEFAULT_LAYOUT_SCALE,
        help="整体缩放，越小越靠近中心",
    )
    parser.add_argument("--font-size", type=int, default=DEFAULT_FONT_SIZE, help="标签字号")
    parser.add_argument(
        "--peripheral-pull",
        type=float,
        default=DEFAULT_PERIPHERAL_PULL,
        help="外围节点向中心收缩比例，越小越靠近中心",
    )
    parser.add_argument("--title", default="", help="图标题，可留空")
    parser.add_argument("--seed", type=int, default=DEFAULT_LAYOUT_SEED, help="布局随机种子")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    labels, matrix = read_matrix(args.input, args.sheet)
    graph = build_graph(labels, matrix, args.edge_percentile, args.top_edges_per_node)
    draw_graph(
        graph,
        args.output,
        args.title or None,
        args.seed,
        args.layout_k,
        args.layout_scale,
        args.font_size,
        args.peripheral_pull,
    )
    print(f"已输出: {args.output}")


if __name__ == "__main__":
    main()
