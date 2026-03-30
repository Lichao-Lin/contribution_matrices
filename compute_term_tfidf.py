import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
import inflect  # 新增：用于处理单复数

# 初始化复数转换器
p = inflect.engine()

DEFAULT_INPUT = Path(
    r"C:\Users\megamind\Desktop\proquest全文清洗2025.12.16 的定稿(2).xlsm"
)
DEFAULT_OUTPUT = Path(
    r"C:\Users\megamind\Desktop\tf-idf-output.xlsx"
)
DEFAULT_SHEET = "Sheet1"
DEFAULT_COLUMN = 3

TERMS = [
    "thailand",
    "cooperation",
    "tesla",
    "european",
    "tariff",
    "trade",
    "country",
    "international",
    "export",
    "car",
    "energy",
    "development",
    "byd",
    "industry",
    "china",
    "global",
    "price",
    "europe",
    "president",
    "economic",
    "the eu",
    "foreign",
    "support",
    "share",
    "ministry",
    "work",
    "company",
    "government",
    "market",
    "need",
    "new",
    "vehicle",
    "electric",
    "chinese",
    "unfair",
    "technological",
    "cheap",
    "strategic",
    "mutual",
    "local",
    "sustainable",
    "japanese",
    "strong",
    "massive",
]

TOKEN_RE = re.compile(r"[a-z]+(?:'[a-z]+)?")

# 新增函数：获取词的所有形式（单数、复数）
def get_all_forms(term: str) -> list[str]:
    """获取一个词的所有形式（单数、复数、小写）"""
    term_lower = term.lower()
    forms = {term_lower}
    
    # 添加复数形式
    plural = p.plural(term_lower)
    if plural != term_lower:
        forms.add(plural)
    
    # 添加单数形式（如果原词是复数）
    singular = p.singular_noun(term_lower)
    if singular:
        forms.add(singular)
    
    return list(forms)

@dataclass
class TermStats:
    term: str
    count: int
    document_count: int
    tf_idf: float


def normalize_text(value) -> str:
    """归一化文本：转小写，去除所有格's"""
    text = str(value or "").strip().lower()
    # 去除所有格 's 和 s'（如 "country's" -> "country"）
    text = re.sub(r"'s\b", "", text)
    text = re.sub(r"s'\b", "s", text)
    return text


def build_pattern(term: str) -> re.Pattern[str]:
    """构建匹配模式，匹配单数和复数形式，忽略大小写"""
    forms = get_all_forms(term)
    # 按长度降序排序，避免部分匹配
    forms.sort(key=len, reverse=True)
    pattern = "|".join(rf"\b{re.escape(form)}\b" for form in forms)
    return re.compile(pattern, re.IGNORECASE)


def read_documents(file_path: Path, sheet_name: str, column_index: int) -> list[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    documents = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if len(row) < column_index:
            continue
        text = normalize_text(row[column_index - 1])
        if text:
            documents.append(text)

    return documents


def count_total_tokens(documents: list[str]) -> int:
    return sum(len(TOKEN_RE.findall(document)) for document in documents)


def calculate_tfidf(documents: list[str], terms: list[str]) -> list[TermStats]:
    total_documents = len(documents)
    total_tokens = count_total_tokens(documents)

    if total_documents == 0:
        raise ValueError("未读取到任何文档内容，请检查 sheet 名称或列号。")

    if total_tokens == 0:
        raise ValueError("第三列没有可用于分词的英文 token。")

    results = []
    for term in terms:
        pattern = build_pattern(term)
        count = sum(len(pattern.findall(document)) for document in documents)
        document_count = sum(1 for document in documents if pattern.search(document))

        if count == 0 or document_count == 0:
            tf_idf = 0.0
        else:
            tf = count / total_tokens
            idf = math.log(total_documents / document_count)
            tf_idf = tf * idf

        results.append(
            TermStats(
                term=term,
                count=count,
                document_count=document_count,
                tf_idf=tf_idf,
            )
        )

    return results


def write_output(output_path: Path, results: list[TermStats]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "sheet1"
    worksheet.append(["单词", "次数", "条数", "tf-idf"])

    for item in results:
        worksheet.append([item.term, item.count, item.document_count, item.tf_idf])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="统计指定词在 Excel 第三列中的 corpus-level tf-idf")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="输入 Excel 文件路径")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出 Excel 文件路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="工作表名称")
    parser.add_argument("--column", type=int, default=DEFAULT_COLUMN, help="目标列号，从 1 开始")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    documents = read_documents(args.input, args.sheet, args.column)
    results = calculate_tfidf(documents, TERMS)
    write_output(args.output, results)
    print(f"已输出: {args.output}")


if __name__ == "__main__":
    main()
