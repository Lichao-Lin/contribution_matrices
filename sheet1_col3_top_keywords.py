import argparse
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path(
    r"C:\Users\megamind\Desktop\proquest全文清洗2025.12.16 的定稿(2).xlsm"
)
DEFAULT_SHEET = "Sheet1"
DEFAULT_COLUMN = 3
DEFAULT_TOP_N = 43


FALLBACK_STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "been",
    "being",
    "by",
    "for",
    "from",
    "has",
    "have",
    "he",
    "her",
    "hers",
    "him",
    "his",
    "i",
    "in",
    "is",
    "it",
    "its",
    "of",
    "on",
    "or",
    "our",
    "ours",
    "she",
    "that",
    "the",
    "their",
    "theirs",
    "them",
    "they",
    "this",
    "those",
    "to",
    "was",
    "were",
    "will",
    "with",
    "you",
    "your",
    "yours",
    "we",
    "us",
    "not",
    "no",
    "do",
    "does",
    "did",
    "but",
    "if",
    "then",
    "than",
    "so",
    "such",
    "into",
    "over",
    "under",
    "out",
    "up",
    "down",
    "about",
    "after",
    "before",
    "during",
    "because",
    "while",
    "can",
    "could",
    "should",
    "would",
    "may",
    "might",
    "must",
}

EXTRA_STOPWORDS = {
    "china",
    "chinese",
    "said",
    "says",
    "say",
    "also",
    "one",
    "two",
    "new",
    "year",
    "years",
    "today",
    "yesterday",
    "reported",
    "according",
    "mr",
    "mrs",
    "ms",
    "dr",
    "reuters",
    "afp",
    "times",
    "daily",
}


def load_stopwords():
    try:
        from sklearn.feature_extraction.text import ENGLISH_STOP_WORDS

        return set(ENGLISH_STOP_WORDS) | EXTRA_STOPWORDS
    except Exception:
        return FALLBACK_STOPWORDS | EXTRA_STOPWORDS


def clean_text(text: str) -> str:
    text = (text or "").strip().lower()
    lines = text.splitlines()

    # The third column repeats the title in the first line, so skip it.
    if len(lines) > 1:
        text = "\n".join(lines[1:]).strip()

    return text


def tokenize(text: str, stopwords: set[str]) -> list[str]:
    words = re.findall(r"[a-z][a-z']{2,}", text)
    return [word for word in words if word not in stopwords]


def top_keywords(file_path: Path, sheet_name: str, column_index: int, top_n: int):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    stopwords = load_stopwords()
    counter = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < column_index:
            continue
        text = clean_text(row[column_index - 1])
        counter.update(tokenize(text, stopwords))

    return counter.most_common(top_n)


def main():
    parser = argparse.ArgumentParser(description="统计 Sheet1 第三列的高频英文关键词")
    parser.add_argument(
        "top_positional",
        nargs="?",
        type=int,
        help="直接写前多少个关键词，例如: python sheet1_col3_top_keywords.py 100",
    )
    parser.add_argument("--file", default=DEFAULT_INPUT, help="Excel 文件路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="工作表名称")
    parser.add_argument("--column", type=int, default=DEFAULT_COLUMN, help="列号，从 1 开始")
    parser.add_argument("--top", type=int, default=DEFAULT_TOP_N, help="输出前 N 个关键词")
    args = parser.parse_args()

    top_n = args.top_positional if args.top_positional is not None else args.top
    results = top_keywords(Path(args.file), args.sheet, args.column, top_n)

    print("rank\tkeyword\tcount")
    for index, (word, count) in enumerate(results, start=1):
        print(f"{index}\t{word}\t{count}")


if __name__ == "__main__":
    main()
