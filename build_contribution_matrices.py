import argparse
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = Path(
    r"C:\Users\megamind\Desktop\proquest全文清洗2025.12.16 的定稿(2).xlsm"
)
DEFAULT_TERMS = Path(
    r"C:\Users\megamind\Desktop\tf-idf-output.xlsx"
)
DEFAULT_OUTPUT = Path(
    r"C:\Users\megamind\Desktop\contribution-matrices.xlsx"
)
DEFAULT_SHEET = "Sheet1"
DEFAULT_COLUMN = 3

TOKEN_RE = re.compile(r"[a-z]+(?:'[a-z]+)?")
PARAGRAPH_SPLIT_RE = re.compile(r"\n\s*\n+")


@dataclass
class TermInfo:
    term: str
    total_count: int
    document_count: int


def normalize_text(value) -> str:
    return str(value or "").strip().lower()


def build_pattern(term: str) -> re.Pattern[str]:
    return re.compile(rf"(?<![a-z]){re.escape(term.lower())}(?![a-z])")


def load_terms(file_path: Path) -> list[TermInfo]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    terms = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        term = normalize_text(row[0] if row else "")
        if term:
            total_count = int(row[1] or 0)
            document_count = int(row[2] or 0)
            terms.append(
                TermInfo(
                    term=term,
                    total_count=total_count,
                    document_count=document_count,
                )
            )

    if not terms:
        raise ValueError("词表为空，请检查 tf-idf-output.xlsx 的第一列。")

    return terms


def read_paragraphs(file_path: Path, sheet_name: str, column_index: int) -> list[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    paragraphs = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if len(row) < column_index:
            continue

        text = normalize_text(row[column_index - 1])
        if not text:
            continue

        for paragraph in PARAGRAPH_SPLIT_RE.split(text):
            paragraph = paragraph.strip()
            if paragraph:
                paragraphs.append(paragraph)

    if not paragraphs:
        raise ValueError("没有读取到任何段落，请检查输入文件内容。")

    return paragraphs


def count_terms_in_paragraph(paragraph: str, patterns: dict[str, re.Pattern[str]]) -> dict[str, int]:
    counts = {}
    for term, pattern in patterns.items():
        term_count = len(pattern.findall(paragraph))
        if term_count:
            counts[term] = term_count
    return counts


def build_matrices(term_infos: list[TermInfo], paragraphs: list[str]) -> tuple[list[list[int]], list[list[int]]]:
    terms = [item.term for item in term_infos]
    term_index = {term: idx for idx, term in enumerate(terms)}
    patterns = {term: build_pattern(term) for term in terms}
    size = len(terms)

    contribution = [[0 for _ in range(size)] for _ in range(size)]
    weighted = [[0 for _ in range(size)] for _ in range(size)]

    for idx, info in enumerate(term_infos):
        contribution[idx][idx] = info.document_count
        weighted[idx][idx] = info.total_count

    for paragraph in paragraphs:
        counts = count_terms_in_paragraph(paragraph, patterns)
        present_terms = list(counts.keys())

        for i, term_i in enumerate(present_terms):
            idx_i = term_index[term_i]
            count_i = counts[term_i]

            for term_j in present_terms[i + 1 :]:
                idx_j = term_index[term_j]
                count_j = counts[term_j]

                contribution_value = 1
                weighted_value = count_i * count_j

                contribution[idx_i][idx_j] += contribution_value
                contribution[idx_j][idx_i] += contribution_value

                weighted[idx_i][idx_j] += weighted_value
                weighted[idx_j][idx_i] += weighted_value

    return contribution, weighted


def write_matrix_sheet(worksheet, sheet_name: str, terms: list[str], matrix: list[list[int]]) -> None:
    worksheet.title = sheet_name
    worksheet.append(["单词", *terms])

    for term, row in zip(terms, matrix):
        worksheet.append([term, *row])


def write_output(output_path: Path, terms: list[str], contribution: list[list[int]], weighted: list[list[int]]) -> None:
    workbook = Workbook()
    contribution_sheet = workbook.active
    write_matrix_sheet(contribution_sheet, "contribution_matrix", terms, contribution)

    weighted_sheet = workbook.create_sheet("weighted_contribution_matrix")
    write_matrix_sheet(weighted_sheet, "weighted_contribution_matrix", terms, weighted)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据段落共现生成贡献矩阵和加权贡献矩阵")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="输入 xlsm 文件路径")
    parser.add_argument("--terms", type=Path, default=DEFAULT_TERMS, help="词表 Excel 文件路径")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出 xlsx 文件路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="输入工作表名称")
    parser.add_argument("--column", type=int, default=DEFAULT_COLUMN, help="目标列号，从 1 开始")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    term_infos = load_terms(args.terms)
    terms = [item.term for item in term_infos]
    paragraphs = read_paragraphs(args.input, args.sheet, args.column)
    contribution, weighted = build_matrices(term_infos, paragraphs)
    write_output(args.output, terms, contribution, weighted)
    print(f"已输出: {args.output}")


if __name__ == "__main__":
    main()
